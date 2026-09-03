"""Tests for tournament_scheduler.pipeline.stage4_export."""

import hashlib
import logging
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
import json
import re

import openpyxl
import pytest

from tournament_scheduler.html.html_exporter import HtmlExporter
from tournament_scheduler.models import Game, Roster, SeasonPlan, Team, Tournament
from tournament_scheduler.pipeline.cache_manager import ScrapedDataCache
from tournament_scheduler.pipeline.not_started import NOT_STARTED_MESSAGE
from tournament_scheduler.pipeline.pages_bundle import build_public_bundle
from tournament_scheduler.pipeline.pages_publish import bundle_fingerprint
from tournament_scheduler.pipeline.stage4_export import Stage4Error, _dict_to_plan, run
from tournament_scheduler.pipeline.state import PipelineState, StageName, StageStatus


def _write_input_workbook(path: Path, raw: dict) -> None:
    wb = openpyxl.Workbook()
    settings = wb.active
    settings.title = "Innstillinger"
    settings.append(["felt", "verdi"])
    for key in ("start_date", "end_date", "target_tournament_count"):
        if key in raw:
            settings.append([key, raw[key]])

    if "age_groups" in raw:
        age_groups = wb.create_sheet("Aldersgrupper")
        age_groups.append(["age_group", "parallel_games", "round_length_minutes"])
        for age_group in raw["age_groups"]:
            age_groups.append([age_group, raw.get("parallel_games", {}).get(age_group), None])

    teams = wb.create_sheet("Lag")
    teams.append(["club", "label", "age_group"])
    for team in raw.get("teams", []):
        teams.append([team.get("club"), team.get("label"), team.get("age_group")])

    sources = wb.create_sheet("Kilder")
    sources.append(["name", "type", "url"])
    for source in raw.get("sources", []):
        sources.append([source.get("name"), source.get("type"), source.get("url")])

    wb.save(path)


def _file_hashes(root: Path) -> dict[str, str]:
    return {
        path.relative_to(root).as_posix(): hashlib.sha256(path.read_bytes()).hexdigest()
        for path in sorted(p for p in root.rglob("*") if p.is_file())
    }


def _make_plan_dict():
    """Build a minimal but valid plan checkpoint dict."""
    t1 = {
        "date": "2025-10-05",
        "arena": "Kongsberghallen",
        "age_group": "U10",
        "host_club": "Kongsberg",
        "teams": [
            {"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"},
            {"club": "Skien",     "label": "Skien U10A",     "age_group": "U10"},
        ],
        "games": [
            {"home": "Kongsberg U10A", "away": "Skien U10A", "parallel_slot": 0, "round_number": 3},
        ],
        "start_time": "09:00",
    }
    return {
        "plan": {
            "start_date": "2025-09-01",
            "end_date": "2025-12-01",
            "diversity_score": 1.0,
            "pairwise_matchup_score": 1.0,
            "month_balance_score": 1.0,
            "arena_counts": {"Kongsberghallen": 1},
            "manual_adjustments": {
                "locked_dates": ["2025-10-05"],
                "banned_dates": ["2025-11-01"],
            },
            "fairness_gate": {
                "status": "pass",
                "score": 100,
                "metrics": [
                    {"label": "Kamper per lag", "value": 0, "threshold": 2, "status": "pass", "score": 100, "unit": "", "detail": "Lik kampfordeling."},
                    {
                        "key": "hosting_deviation",
                        "label": "Hjemmebanebelastning",
                        "value": 0.0,
                        "threshold": 1,
                        "status": "pass",
                        "score": 100,
                        "unit": "",
                        "detail": "Aldersgruppevis fordeling av hjemmeturneringer: U10 Kongsberg 1 vs ~1.0.",
                        "age_group_breakdown": [
                            {"age_group": "U10", "club": "Kongsberg", "actual": 1, "expected": 1.0},
                            {"age_group": "U10", "club": "Skien", "actual": 0, "expected": 0.0},
                        ],
                    },
                    {"label": "Månedsbalanse", "value": 1.0, "threshold": 0.75, "status": "pass", "score": 100, "unit": "", "detail": "Jevn sesongbelastning."},
                ],
            },
            "tournaments": [t1],
        },
        "llm_confidence": 0.9,
        "llm_reasoning": "great",
        "attempts": 1,
        "llm_skipped": True,
    }


def _make_multi_age_group_plan_dict():
    data = _make_plan_dict()
    data["plan"]["tournaments"].append(
        {
            "date": "2025-11-02",
            "arena": "Bærum ishall",
            "age_group": "JU11",
            "host_club": "Jutul",
            "teams": [
                {"club": "Jutul", "label": "Jutul JU11A", "age_group": "JU11"},
                {"club": "Holmen", "label": "Holmen JU11A", "age_group": "JU11"},
            ],
            "games": [
                {"home": "Jutul JU11A", "away": "Holmen JU11A", "parallel_slot": 0, "round_number": 1},
            ],
        }
    )
    data["plan"]["arena_counts"]["Bærum ishall"] = 1
    return data


def _make_spond_plan_dict():
    data = _make_plan_dict()
    data["plan"]["tournaments"][0]["teams"] = [
        {"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"},
        {"club": "Skien", "label": "Skien U10A", "age_group": "U10"},
        {"club": "Holmen", "label": "Holmen U10A", "age_group": "U10"},
    ]
    data["plan"]["tournaments"][0]["games"] = [
        {"home": "Kongsberg U10A", "away": "Skien U10A", "parallel_slot": 0, "round_number": 1},
        {"home": "Kongsberg U10A", "away": "Holmen U10A", "parallel_slot": 0, "round_number": 2},
        {"home": "Skien U10A", "away": "Holmen U10A", "parallel_slot": 0, "round_number": 3},
    ]
    return data


class TestDictToPlan:
    def test_round_trips_plan(self):
        plan_dict = _make_plan_dict()["plan"]
        plan = _dict_to_plan(plan_dict)
        assert isinstance(plan, SeasonPlan)
        assert len(plan.tournaments) == 1
        t = plan.tournaments[0]
        assert t.arena == "Kongsberghallen"
        assert t.start_time == "09:00"
        assert len(t.games) == 1
        assert t.games[0].home.label == "Kongsberg U10A"
        assert t.games[0].round_number == 3
        assert plan.manual_adjustments["locked_dates"] == ["2025-10-05"]

    def test_round_trips_manual_booking_reason(self):
        plan_dict = _make_plan_dict()["plan"]
        plan_dict["tournaments"][0]["manual_booking_reason"] = "Kalender utilgjengelig for Tønsberg — istid må bookes manuelt."
        plan = _dict_to_plan(plan_dict)
        assert plan.tournaments[0].manual_booking_reason == (
            "Kalender utilgjengelig for Tønsberg — istid må bookes manuelt."
        )

    def test_round_trips_arena_day_collisions(self):
        plan_dict = _make_plan_dict()["plan"]
        plan_dict["arena_day_collisions"] = [
            {
                "date": "2025-10-05",
                "arena": "Jarahallen",
                "age_group": "U7",
                "host_club": "Jar",
                "conflicting_age_group": "U10",
                "conflicting_host_club": "Jar",
                "reason": "same_arena_same_day",
            }
        ]
        plan = _dict_to_plan(plan_dict)
        assert plan.arena_day_collisions[0]["arena"] == "Jarahallen"
        assert plan.arena_day_collisions[0]["conflicting_age_group"] == "U10"

    def test_handles_missing_dates(self):
        plan_dict = {"tournaments": [], "diversity_score": 0.0,
                     "pairwise_matchup_score": 0.0, "month_balance_score": 0.0,
                     "arena_counts": {}}
        plan = _dict_to_plan(plan_dict)
        assert plan.start_date is None

    def test_raises_on_missing_tournament_date(self):
        plan_dict = _make_plan_dict()["plan"]
        plan_dict["tournaments"][0].pop("date")
        with pytest.raises(ValueError, match="date"):
            _dict_to_plan(plan_dict)

    def test_raises_on_empty_tournament_date(self):
        plan_dict = _make_plan_dict()["plan"]
        plan_dict["tournaments"][0]["date"] = ""
        with pytest.raises(ValueError, match="date"):
            _dict_to_plan(plan_dict)

    def test_warns_when_dropping_game_with_unknown_team_label(self, caplog):
        plan_dict = _make_plan_dict()["plan"]
        plan_dict["tournaments"][0]["teams"].append(
            {"club": "Holmen", "label": "Holmen U10A", "age_group": "U10"}
        )
        plan_dict["tournaments"][0]["games"].append(
            {"home": "Kongsberg U10A", "away": "Missing U10A", "parallel_slot": 0, "round_number": 4}
        )

        with caplog.at_level(logging.WARNING, logger="tournament_scheduler.pipeline.stage4_helpers"):
            plan = _dict_to_plan(plan_dict)

        assert len(plan.tournaments[0].games) == 1
        assert any("Missing U10A" in record.message for record in caplog.records)


class TestRunStage4:
    def test_export_warns_and_lists_colliding_tournaments_instead_of_blocking(self, tmp_path):
        """Arena/sequence collisions no longer block the full export: they are
        surfaced as a warning, recorded in the checkpoint, and listed in a
        dedicated manual-schedule view (manual_schedule.html)."""
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(StageName.CONFIG, {"round_length_minutes": {"U10": 15}}, status=StageStatus.DONE)
        plan_checkpoint = _make_plan_dict()
        first = plan_checkpoint["plan"]["tournaments"][0]
        first["id"] = "first"
        first["start_time"] = "09:00"
        plan_checkpoint["plan"]["tournaments"].append(
            {
                **first,
                "id": "second",
                "start_time": "09:30",
            }
        )

        result = run(
            plan_checkpoint,
            state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )

        files = result.get("output_files", {})
        assert "excel" in files
        assert "html" in files
        assert "html_report" in files
        assert "manual_schedule" in files
        assert Path(files["manual_schedule"]).exists()
        assert result["arena_day_collisions"], "checkpoint should record remaining collisions"
        assert "Kongsberghallen" in result["arena_day_collisions"][0]["message"]

        manual_html = Path(files["manual_schedule"]).read_text(encoding="utf-8")
        assert "Må planlegges manuelt" in manual_html
        assert "Kongsberghallen" in manual_html

        # The season-plan page must link to the manual view in its navbar.
        html = Path(files["html"]).read_text(encoding="utf-8")
        assert 'href="manual_schedule.html"' in html

        # The report must present the plan as usable-but-needs-manual-followup.
        report_html = Path(files["html_report"]).read_text(encoding="utf-8")
        assert "MÅ SJEKKES" in report_html
        assert "Manuell istidplanlegging" in report_html

    def test_produces_excel_file(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(StageName.CONFIG, {"round_length_minutes": {"U10": 15}}, status=StageStatus.DONE)
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        files = result.get("output_files", {})
        assert "excel" in files
        assert Path(files["excel"]).exists()
        workbook = openpyxl.load_workbook(files["excel"])
        overview = workbook["Sesongoversikt"]
        rows = list(overview.iter_rows(values_only=True))
        assert rows[1][7] == "09:00"
        assert rows[1][8] == "09:45"

    def test_not_started_plan_writes_placeholder_outputs(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        input_file = tmp_path / "input.xlsx"
        _write_input_workbook(input_file, {
            "start_date": "2025-09-01",
            "end_date": "2025-12-01",
            "teams": [],
        })
        state.write_stage(StageName.CONFIG, {"input_path": str(input_file), "teams": []}, status=StageStatus.DONE)
        checkpoint = {
            "not_started": True,
            "plan": {
                "start_date": "2025-09-01",
                "end_date": "2025-12-01",
                "tournaments": [],
                "placeholder": "not_started",
                "message": NOT_STARTED_MESSAGE,
            },
        }

        result = run(checkpoint, state, export_dir=str(tmp_path / "export"), timestamped_export=False)

        files = result["output_files"]
        assert result["not_started"] is True
        assert set(["excel", "ical", "csv_games", "csv_overview", "input_html", "calendars_html", "html", "html_report", "spond", "spond_games", "review_packets"]) <= set(files)
        html = Path(files["html"]).read_text(encoding="utf-8")
        assert "Ikke begynt" in html
        assert "<meta name=\"viewport\"" in html
        assert "<div class=\"icon\">🏒</div>" in html
        assert "Ikke begynt" in Path(files["csv_games"]).read_text(encoding="utf-8")
        workbook = openpyxl.load_workbook(files["excel"])
        assert workbook.active["A1"].value.startswith("Ikke begynt")

    def test_includes_fairness_gate_sheet(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        files = result.get("output_files", {})
        workbook = openpyxl.load_workbook(files["excel"])
        assert "Rettferdighetskontroll" in workbook.sheetnames
        assert "Rettferdighetsjusteringer" in workbook.sheetnames
        old_gate_label = "Rettferdighets" + "gate"
        old_adjustment_sheet = "Fairness" + "justeringer"
        assert old_gate_label not in workbook.sheetnames
        assert old_adjustment_sheet not in workbook.sheetnames
        sheet = workbook["Rettferdighetskontroll"]
        rows = list(sheet.iter_rows(values_only=True))
        assert rows[0][0] == "Overordnet status"
        assert rows[2][0] == "Metrikk"
        assert rows[3][0] == "Kamper per lag"
        adj = workbook["Rettferdighetsjusteringer"]
        adj_rows = list(adj.iter_rows(values_only=True))
        assert adj_rows[0][0] == "Rettferdighetsjusteringer per lag"
        assert "fairness" not in str(adj_rows[1][0]).lower()
        assert adj_rows[3][0] == "Lag"

    def test_generates_html_with_configured_age_group_filters(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_input_workbook(
            input_path,
            {
                "start_date": "2025-09-01",
                "end_date": "2025-12-01",
                "age_groups": ["U10", "U11", "U12", "JU11"],
                "parallel_games": {"U10": 2, "U11": 2, "U12": 2, "JU11": 2},
                "teams": [],
                "sources": [],
            },
        )
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(
            StageName.CONFIG,
            {"input_path": str(input_path), "round_length_minutes": {}},
            status=StageStatus.DONE,
        )
        result = run(
            _make_multi_age_group_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        files = result.get("output_files", {})
        html_path = Path(files["html"])
        report_path = Path(files["html_report"])
        html = html_path.read_text(encoding="utf-8")
        report_html = report_path.read_text(encoding="utf-8")

        assert '<option value="U10">U10</option>' in html
        assert '<option value="JU11">JU11</option>' in html
        assert '<option value="U12">U12</option>' in html
        assert 'Alle (U10 + U11 + U12 + JU11)' in html
        assert 'id="themeToggle"' in html
        assert 'class="theme-toggle"' in html
        assert 'href="season_plan.xlsx"' in html
        assert 'href="season_plan.csv"' in html
        assert 'href="season_plan.ics"' in html
        assert 'href="season_plan.csv" class="export-link-btn"' in html or 'href="season_plan.csv"' in html
        assert html.index('class="export-links"') < html.index('class="header-main"')
        assert report_html.index('class="export-links"') < report_html.index('class="header-main"')
        assert html.index('class="export-links"') < html.index('class="stat-badge"')
        assert report_html.index('class="export-links"') < report_html.index('class="stat-badge"')
        old_gate_label = 'Rettferdighets' + 'gate'
        old_adjustment_label = 'Fairness' + '-justeringer'
        assert old_gate_label not in html
        assert old_adjustment_label not in html
        assert 'Rettferdighetskontroll' not in html
        assert 'Rettferdighetsjusteringer' not in html
        assert 'Kvalitetsgjennomgang' not in html
        assert 'id="timeline"' in html
        assert 'class="filters"' in html
        assert 'class="count-bar"' in html
        assert 'Ser planen jevn ut?' in report_html
        assert 'Rettferdighetsjusteringer' in report_html
        assert 'Per aldersgruppe og klubb: faktisk vs forventet hjemmeturneringer' in report_html
        assert 'Aldersgruppevis fordeling av hjemmeturneringer: U10 Kongsberg 1 vs ~1.0.' in report_html
        assert 'id="reportOverview"' in report_html
        assert 'Kan planen brukes?' in report_html
        assert 'Hva må sjekkes eller endres?' in report_html
        assert 'Ja — planen kan brukes' in report_html or 'Nei — planen bør stoppes' in report_html
        assert 'Hva skjer per aldersgruppe?' in report_html
        assert 'Hva må hver klubb vurdere?' in report_html
        assert 'Turneringer som skal gjennomgås' in report_html
        assert 'Detaljerte måltall og kontroller' in report_html
        assert 'Klubben bør sjekke' in report_html
        assert 'Vis hvorfor' in report_html
        assert 'report-status-pill' in report_html
        assert report_html.index('id="reportOverview"') < report_html.index('Ser planen jevn ut?')
        assert report_html.index('Hva må sjekkes eller endres?') < report_html.index('Detaljerte måltall og kontroller')
        assert report_html.index('Ser planen jevn ut?') < report_html.index('Detaljerte måltall og kontroller')
        assert report_html.index('id="reportOverview"') < report_html.index('Vis hvorfor')
        assert old_gate_label not in report_html
        assert old_adjustment_label not in report_html
        assert 'Rådgivende kontroll' in report_html
        assert 'Manglende klubber' in report_html
        assert 'id="clubReviewSummary"' in report_html
        assert 'id="teamStats"' in report_html
        assert 'id="travelStats"' in report_html
        assert 'id="heatmapSection"' in report_html
        # Heatmap must appear immediately after the hero block — before the card-grid and all other sections.
        assert report_html.index('id="heatmapSection"') < report_html.index('class="report-card-grid"')
        assert report_html.index('id="heatmapSection"') < report_html.index('id="priorityActions"')
        assert 'id="clubDashboard"' not in report_html
        assert 'style="display:none' not in report_html
        assert 'id="timeline"' not in report_html
        assert 'class="filters"' not in report_html
        assert 'class="count-bar"' not in report_html
        assert 'Klassetrinn' not in report_html
        assert 'Nullstill filter' not in report_html
        assert 'Viser 80 av 80 turneringer' not in report_html
        schedule_script = re.search(r"<script>\n(.*?)\n</script>", html, re.S).group(1)
        report_script = re.search(r"<script>\n(.*?)\n</script>", report_html, re.S).group(1)
        for schedule_identifier in ('filterAge', 'timeline', 'buildMatchHTML', 'function render()'):
            assert schedule_identifier in schedule_script
            assert schedule_identifier not in report_script
        assert 'rvv-theme' in report_script
        assert 'HEATMAP' in report_script
        assert 'clubSummaryBody' in report_script
        assert 'clubDashName' not in report_script
        assert 'debug-dashboard' not in html.lower()
        assert not re.search(r"[\U0001F300-\U0001FAFF]", html)
        assert not re.search(r"[\U0001F300-\U0001FAFF]", report_html)
        # Hero verdict pill should be present and use a plain-language decision.
        assert 'report-status-pill' in report_html
        assert 'KAN BRUKES' in report_html or 'BLOKKER' in report_html

    def test_review_summary_collapses_when_it_only_repeats_main_assessment(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            _make_plan_dict(),
            state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")

        assert 'review-summary-panel--compact' in report_html
        assert 'Kortversjon av kontrollen' in report_html

    def test_report_missing_hosts_uses_canonical_club_aliases(self, tmp_path):
        """Short RVV club aliases should not trigger false missing-host warnings."""
        rvv_hosts = [
            "Ringerike",
            "Tønsberg",
            "Frisk Asker",
            "Sandefjord",  # Alias for canonical Sandefjord Penguins.
            "Jar",
            "Holmen",
            "Skien",
            "Jutul",
            "Kongsberg",
        ]
        start = date(2025, 10, 5)
        tournaments = []
        for index, host in enumerate(rvv_hosts):
            opponent = "Skien" if host != "Skien" else "Kongsberg"
            tournaments.append(
                {
                    "date": (start + timedelta(days=index * 7)).isoformat(),
                    "arena": f"{host} arena",
                    "age_group": "U10",
                    "host_club": host,
                    "teams": [
                        {"club": host, "label": f"{host} U10A", "age_group": "U10"},
                        {"club": opponent, "label": f"{opponent} U10A", "age_group": "U10"},
                    ],
                    "games": [
                        {
                            "home": f"{host} U10A",
                            "away": f"{opponent} U10A",
                            "parallel_slot": 0,
                            "round_number": 1,
                        },
                    ],
                }
            )
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            {
                "plan": {
                    "start_date": "2025-10-01",
                    "end_date": "2025-12-31",
                    "diversity_score": 1.0,
                    "pairwise_matchup_score": 1.0,
                    "month_balance_score": 1.0,
                    "arena_counts": {},
                    "tournaments": tournaments,
                }
            },
            state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")

        assert "Rådgivende kontroll" in report_html
        assert "Alle 9 RVV-klubber har minst én hjemmeturnering." in report_html
        assert "Følgende RVV-klubber har ingen hjemmeturnering" not in report_html

    def test_html_filters_fall_back_to_plan_age_groups_when_input_omits_them(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_input_workbook(
            input_path,
            {
                "start_date": "2025-09-01",
                "end_date": "2025-12-01",
                "teams": [
                    {"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"},
                ],
                "parallel_games": {"U10": 2},
                "sources": [],
            },
        )
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(
            StageName.CONFIG,
            {"input_path": str(input_path), "round_length_minutes": {}},
            status=StageStatus.DONE,
        )
        result = run(
            _make_multi_age_group_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        html = Path(result["output_files"]["html"]).read_text(encoding="utf-8")

        assert '<option value="U10">U10</option>' in html
        assert '<option value="JU11">JU11</option>' in html
        assert 'Alle (JU11 + U10)' in html or 'Alle (U10 + JU11)' in html

    def test_html_tournament_details_group_matches_by_round(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        files = result.get("output_files", {})
        html = Path(files["html"]).read_text(encoding="utf-8")

        payload = json.loads(HtmlExporter._plan_to_json(_dict_to_plan(_make_plan_dict()["plan"])))
        assert payload[0]["m"][0][3] == 3
        assert 'Kamper per runde' in html
        assert 'round-group-header' in html
        assert 'Runde ' in html

    def test_produces_ical_file(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(StageName.CONFIG, {"round_length_minutes": {"U10": 15}}, status=StageStatus.DONE)
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        files = result.get("output_files", {})
        assert "ical" in files
        ics_path = Path(files["ical"])
        assert ics_path.exists()
        assert ics_path.suffix == ".ics"
        content = ics_path.read_text()
        assert "BEGIN:VCALENDAR" in content
        assert "VEVENT" in content
        assert "DTSTART:20251005T090000Z" in content
        assert "DTEND:20251005T094500Z" in content

    def test_writes_timestamped_exports_without_flat_copies(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=True,
        )
        files = result.get("output_files", {})

        timestamped_paths = [Path(files[key]) for key in ("excel", "ical", "csv_games", "csv_overview", "html", "html_report", "spond", "spond_games")]
        timestamp_dirs = {path.parent for path in timestamped_paths}
        assert len(timestamp_dirs) == 1
        timestamp_dir = timestamp_dirs.pop()
        assert timestamp_dir.parent == tmp_path / "export"
        assert timestamp_dir.name

        for path in timestamped_paths:
            assert path.exists()
            assert path.parent == timestamp_dir

        assert not any(key.endswith("_flat") for key in files)
        assert list((tmp_path / "export").glob("*.xlsx")) == []
        assert list((tmp_path / "export").glob("*.csv")) == []
        assert list((tmp_path / "export").glob("*.ics")) == []
        assert list((tmp_path / "export").glob("*.html")) == []

    def test_default_run_produces_timestamped_subfolder(self, tmp_path):
        """run() with no explicit timestamped_export should default to True and write into a timestamped subfolder."""
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
        )
        files = result.get("output_files", {})

        excel_path = Path(files["excel"])
        export_dir = tmp_path / "export"
        # The file must be inside a subfolder of export/, not directly in export/
        assert excel_path.parent != export_dir, "Expected a timestamped subfolder, not flat export"
        assert excel_path.parent.parent == export_dir
        # Subfolder name must match YYYY-MM-DDTHHMM pattern
        assert re.match(r"^\d{4}-\d{2}-\d{2}T\d{4}$", excel_path.parent.name), (
            f"Expected timestamped subfolder name like 2025-10-05T0900, got {excel_path.parent.name!r}"
        )

    def test_stage4_spond_export_uses_tournament_rows(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(StageName.CONFIG, {"round_length_minutes": {"U10": 15}}, status=StageStatus.DONE)
        result = run(
            _make_spond_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        files = result.get("output_files", {})
        workbook = openpyxl.load_workbook(files["spond"])
        sheet = workbook["Sesongplan"]
        rows = list(sheet.iter_rows(values_only=True))

        assert rows[0][0:5] == ("Dato", "Aktivitet", "Sted", "Start", "Slutt")
        assert rows[1][3] == "09:00"
        assert rows[1][4] == "09:45"
        assert rows[1][9] == "turnering"
        assert len(rows) == 2  # header + one tournament row, not one row per game

        attachment = openpyxl.load_workbook(files["spond_games"])
        assert len(attachment.sheetnames) == 1
        attachment_rows = list(attachment[attachment.sheetnames[0]].iter_rows(values_only=True))
        header_row = next(i for i, row in enumerate(attachment_rows) if row[:4] == ("Runde", "Hjemmelag", "Bortelag", "Parallellbane"))
        assert attachment_rows[header_row][0:4] == ("Runde", "Hjemmelag", "Bortelag", "Parallellbane")
        assert attachment_rows[header_row + 1][1] == "Kongsberg U10A"

    def test_produces_csv_files(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        files = result.get("output_files", {})
        assert "csv_games" in files
        assert "csv_overview" in files
        games_path = Path(files["csv_games"])
        assert games_path.exists()
        lines = games_path.read_text().splitlines()
        assert lines[0] == "date,arena,age_group,home,away,parallel_slot"
        assert len(lines) > 1  # header + at least one game row

    def test_explicit_build_timestamp_controls_metadata_and_timestamped_directory(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            _make_plan_dict(),
            state,
            export_dir=str(tmp_path / "export"),
            build_timestamp="2025-01-02T03:04:05+00:00",
        )

        assert result["generated_at"] == "2025-01-02T03:04:05+00:00"
        assert Path(result["output_files"]["html"]).parent.name == "2025-01-02T0304"
        envelope = state.read_envelope(StageName.EXPORT)
        assert envelope["data"]["generated_at"] == "2025-01-02T03:04:05+00:00"

    def test_source_date_epoch_controls_stage4_build_timestamp(self, tmp_path, monkeypatch):
        monkeypatch.setenv("SOURCE_DATE_EPOCH", "1735732800")
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            _make_plan_dict(),
            state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )

        assert result["generated_at"] == datetime.fromtimestamp(1735732800, tz=timezone.utc).isoformat()

    def test_same_build_timestamp_produces_identical_export_file_bytes(self, tmp_path):
        build_timestamp = "2025-01-02T03:04:05+00:00"
        export_a = tmp_path / "export-a"
        export_b = tmp_path / "export-b"

        run(
            _make_plan_dict(),
            PipelineState(tmp_path / "pipeline-a"),
            export_dir=str(export_a),
            timestamped_export=False,
            build_timestamp=build_timestamp,
        )
        run(
            _make_plan_dict(),
            PipelineState(tmp_path / "pipeline-b"),
            export_dir=str(export_b),
            timestamped_export=False,
            build_timestamp=build_timestamp,
        )

        hashes_a = _file_hashes(export_a)
        hashes_b = _file_hashes(export_b)
        assert hashes_a == hashes_b
        assert any(name.endswith(".xlsx") for name in hashes_a)

    def test_public_bundle_fingerprint_is_stable_for_unchanged_export_content(self, tmp_path):
        build_timestamp = "2025-01-02T03:04:05+00:00"
        export_a = tmp_path / "export-a"
        export_b = tmp_path / "export-b"
        export_c = tmp_path / "export-c"

        run(
            _make_plan_dict(),
            PipelineState(tmp_path / "pipeline-a"),
            export_dir=str(export_a),
            timestamped_export=False,
            build_timestamp=build_timestamp,
        )
        run(
            _make_plan_dict(),
            PipelineState(tmp_path / "pipeline-b"),
            export_dir=str(export_b),
            timestamped_export=False,
            build_timestamp=build_timestamp,
        )
        changed_plan = _make_plan_dict()
        changed_plan["plan"]["tournaments"][0]["arena"] = "Jarahallen"
        run(
            changed_plan,
            PipelineState(tmp_path / "pipeline-c"),
            export_dir=str(export_c),
            timestamped_export=False,
            build_timestamp=build_timestamp,
        )

        for source, target in (
            (export_a, tmp_path / "public-a"),
            (export_b, tmp_path / "public-b"),
            (export_c, tmp_path / "public-c"),
        ):
            result = build_public_bundle(str(source), str(target))
            assert result.status == "ok"

        assert _file_hashes(tmp_path / "public-a") == _file_hashes(tmp_path / "public-b")
        assert bundle_fingerprint(str(tmp_path / "public-a")) == bundle_fingerprint(str(tmp_path / "public-b"))
        assert bundle_fingerprint(str(tmp_path / "public-a")) != bundle_fingerprint(str(tmp_path / "public-c"))

    def test_export_metadata_includes_generation_timestamp_and_input_path(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_input_workbook(
            input_path,
            {
                "start_date": "2025-09-01",
                "end_date": "2025-12-01",
                "teams": [],
                "sources": [],
            },
        )
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(
            StageName.CONFIG,
            {"input_path": str(input_path), "round_length_minutes": {}},
            status=StageStatus.DONE,
        )
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )

        assert result["generated_at"]
        assert result["input_path"] == str(input_path)
        envelope = state.read_envelope(StageName.EXPORT)
        assert envelope["data"]["generated_at"] == result["generated_at"]
        assert envelope["data"]["input_path"] == str(input_path)

        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")
        assert "Generert" in report_html
        assert input_path.name in report_html
        assert str(input_path) in report_html

    def test_report_renders_generated_at_in_europe_oslo_time(self, tmp_path):
        plan = _dict_to_plan(_make_plan_dict()["plan"])
        export_path = tmp_path / "export" / "season_plan.html"

        HtmlExporter().export(
            plan,
            export_path,
            meta={
                "updated_at": "2025-01-01T12:00:00+00:00",
                "source_count": 1,
                "total_events": 1,
            },
            pipeline_meta={
                "generated_at": "2025-01-01T12:00:00+00:00",
            },
        )

        report_html = export_path.with_name("season_plan_report.html").read_text(encoding="utf-8")
        assert "2025-01-01 13:00 CET" in report_html

    def test_marks_checkpoint_done(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        assert state.is_done(StageName.EXPORT)

    def test_raises_on_missing_plan(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        with pytest.raises(Stage4Error, match="Stage 3"):
            run({}, state, export_dir=str(tmp_path / "export"), strict=True)

    def test_conclusion_injects_weakest_metric_name(self, tmp_path):
        """Conclusion must include weakest metric label when metric_warnings exist."""
        data = _make_plan_dict()
        data["plan"]["fairness_gate"] = {
            "status": "warn",
            "score": 70,
            "metrics": [
                {"label": "Kampbalanse", "value": 3, "threshold": 2, "status": "warn", "score": 60, "unit": "", "detail": "Ujevn fordeling."},
                {"label": "Hjemmebanebelastning", "value": 1.5, "threshold": 1, "status": "fail", "score": 40, "unit": "", "detail": "For stor belastning."},
            ],
        }
        state = PipelineState(tmp_path / "pipeline")
        result = run(data, state, export_dir=str(tmp_path / "export"), timestamped_export=False)
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")
        # Weakest metric is "Hjemmebanebelastning" (status=fail, score=40)
        assert "Svakeste metrikk: Hjemmebanebelastning" in report_html
        # Hero pill must show the issue count: gate warn (1) + 2 metric warnings (2) + missing hosts (1) = 4.
        assert "4 punkt(er)" in report_html

    def test_conclusion_injects_tournament_count(self, tmp_path):
        """Conclusion must include tournament count from active_tournaments."""
        data = _make_plan_dict()
        state = PipelineState(tmp_path / "pipeline")
        result = run(data, state, export_dir=str(tmp_path / "export"), timestamped_export=False)
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")
        # Plan has 1 active tournament; the answer string embeds "1 turneringer"
        assert "1 turneringer" in report_html

    def test_conclusion_injects_month_span(self, tmp_path):
        """Conclusion must include the Norwegian month-span derived from plan dates."""
        data = _make_plan_dict()
        # Plan start_date=2025-09-01, end_date=2025-12-01 → "september–desember"
        state = PipelineState(tmp_path / "pipeline")
        result = run(data, state, export_dir=str(tmp_path / "export"), timestamped_export=False)
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")
        assert "september" in report_html
        assert "desember" in report_html

    def test_conclusion_injects_most_travel_team(self, tmp_path):
        """Conclusion must name the team with the most travel distance."""
        data = _make_plan_dict()
        state = PipelineState(tmp_path / "pipeline")
        result = run(data, state, export_dir=str(tmp_path / "export"), timestamped_export=False)
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")
        # The plan has Kongsberg and Skien; whichever has most travel should appear in conclusion
        assert "Mest reisende lag:" in report_html

    def test_conclusion_injects_weakest_metric_score(self, tmp_path):
        """Conclusion must include the numeric score of the weakest metric."""
        data = _make_plan_dict()
        data["plan"]["fairness_gate"] = {
            "status": "warn",
            "score": 65,
            "metrics": [
                {"label": "Kampbalanse", "value": 3, "threshold": 2, "status": "warn", "score": 55, "unit": "", "detail": "Ujevn fordeling."},
            ],
        }
        state = PipelineState(tmp_path / "pipeline")
        result = run(data, state, export_dir=str(tmp_path / "export"), timestamped_export=False)
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")
        # Score value 55 must appear in the metric score annotation
        assert "Svakeste metrikk: Kampbalanse (55%)" in report_html

    def test_conclusion_injects_fairness_submetric_detail(self, tmp_path):
        """Conclusion must include fairness sub-metric detail for warn/fail status."""
        data = _make_plan_dict()
        data["plan"]["fairness_gate"] = {
            "status": "fail",
            "score": 30,
            "metrics": [
                {"label": "Hjemmebanebelastning", "value": 2.5, "threshold": 1, "status": "fail", "score": 30, "unit": "", "detail": "Kritisk skjevfordeling hjemme."},
            ],
        }
        state = PipelineState(tmp_path / "pipeline")
        result = run(data, state, export_dir=str(tmp_path / "export"), timestamped_export=False)
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")
        # The fairness detail string must appear for fail status
        assert "Fairness-avvik: Hjemmebanebelastning" in report_html
        assert "Kritisk skjevfordeling hjemme." in report_html

    def test_manual_schedule_view_lists_calendarless_host_tournaments(self, tmp_path):
        """Tournaments hosted by a club whose calendar could not be scraped are
        listed in the manual-schedule view (provisional hall time), and the
        report flags the plan as needing manual follow-up."""
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(StageName.CONFIG, {"round_length_minutes": {"U10": 15}}, status=StageStatus.DONE)
        plan_checkpoint = _make_plan_dict()
        first = plan_checkpoint["plan"]["tournaments"][0]
        first["id"] = "tnsb001"
        first["host_club"] = "Tønsberg"
        first["arena"] = "Tønsberghallen"
        first["manual_booking_reason"] = "Kalender utilgjengelig for Tønsberg — istid må bookes manuelt."

        result = run(
            plan_checkpoint,
            state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )

        assert result["manual_booking_count"] == 1
        files = result.get("output_files", {})
        manual_html = Path(files["manual_schedule"]).read_text(encoding="utf-8")
        assert "Kalender utilgjengelig" in manual_html
        assert "Tønsberg" in manual_html
        assert "tnsb001" in manual_html
        html = Path(files["html"]).read_text(encoding="utf-8")
        assert 'href="manual_schedule.html"' in html
        assert "MÅ BOOKES MANUELT" in html
        report_html = Path(files["html_report"]).read_text(encoding="utf-8")
        assert "MÅ SJEKKES" in report_html
        assert "Manuell istidsplanlegging" in report_html

    def test_calendars_html_generated_when_scrape_cache_populated(self, tmp_path):
        """stage4 should write calendars.html when the scrape cache contains events and link it in the navbar."""
        work_dir = tmp_path / "pipeline"
        state = PipelineState(work_dir)
        # Populate scrape cache so generate_html has data to render.
        # total_events and source_count must be at the top level — stage4_export checks those keys directly,
        # not the _meta sub-dict.
        cache = ScrapedDataCache(str(work_dir))
        cache.write({
            "_meta": {
                "updated_at": "2025-01-01T00:00:00",
            },
            "source_count": 1,
            "total_events": 2,
            "sources": {
                "TestClub": {
                    "events": [
                        {"date": "05.10.2025", "title": "Test", "source": "TestClub", "url": ""},
                    ],
                    "scrape_timestamp": "2025-01-01T00:00:00",
                }
            },
        })
        export_dir = tmp_path / "export"
        result = run(_make_plan_dict(), state, export_dir=str(export_dir), timestamped_export=False)
        files = result.get("output_files", {})
        assert "calendars_html" in files
        assert Path(files["calendars_html"]).exists()
        # The navbar in season_plan.html must contain a link to calendars.html
        html = Path(files["html"]).read_text(encoding="utf-8")
        assert 'href="calendars.html"' in html, "season_plan.html navbar must link to calendars.html when scrape cache is present"

    def test_calendars_html_absent_and_nav_link_hidden_when_no_scrape_cache(self, tmp_path):
        """When no scrape cache exists, calendars.html should not be generated and the navbar link should be hidden."""
        state = PipelineState(tmp_path / "pipeline")
        export_dir = tmp_path / "export"
        result = run(_make_plan_dict(), state, export_dir=str(export_dir), timestamped_export=False)
        files = result.get("output_files", {})
        assert "calendars_html" not in files
        calendars_path = export_dir / "calendars.html"
        assert not calendars_path.exists()
        html_path = Path(files["html"])
        html = html_path.read_text(encoding="utf-8")
        assert "Skrapede kalendere" not in html

    def test_input_html_generated_and_linked_when_input_workbook_configured(self, tmp_path):
        """stage4 should write input.html from the Lag sheet and link it from season_plan.html's navbar."""
        input_path = tmp_path / "input.xlsx"
        _write_input_workbook(
            input_path,
            {
                "start_date": "2025-09-01",
                "end_date": "2025-12-01",
                "teams": [
                    {"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"},
                    {"club": "Skien", "label": "Skien U10A", "age_group": "U10"},
                ],
                "sources": [],
            },
        )
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(
            StageName.CONFIG,
            {"input_path": str(input_path), "round_length_minutes": {}},
            status=StageStatus.DONE,
        )
        export_dir = tmp_path / "export"
        result = run(_make_plan_dict(), state, export_dir=str(export_dir), timestamped_export=False)
        files = result.get("output_files", {})

        assert "input_html" in files
        input_html_path = Path(files["input_html"])
        assert input_html_path.exists()
        input_html = input_html_path.read_text(encoding="utf-8")
        assert "Kongsberg U10A" in input_html
        assert "Skien U10A" in input_html

        html = Path(files["html"]).read_text(encoding="utf-8")
        assert 'href="input.html"' in html, "season_plan.html navbar must link to input.html when it was generated"

    def test_activity_artifacts_generated_when_input_workbook_has_activity_sheet(self, tmp_path):
        """stage4 should write activities.json and activities/index.html from the public activity sheet."""
        input_path = tmp_path / "input.xlsx"
        _write_input_workbook(
            input_path,
            {
                "start_date": "2026-01-01",
                "end_date": "2026-12-31",
                "teams": [{"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"}],
                "sources": [{"name": "Hemmelig kilde", "type": "ical", "url": "https://secret.example.com/feed.ics"}],
            },
        )
        wb = openpyxl.load_workbook(input_path)
        activities = wb.create_sheet("Aktiviteter")
        activities.append(["Måned", "Dato", "Aktivitet", "Sted"])
        activities.append(["Januar", 17, "Spillerutvikling JU14", "Sandefjord"])
        wb.save(input_path)

        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(
            StageName.CONFIG,
            {"input_path": str(input_path), "round_length_minutes": {}, "start_date": "2026-01-01"},
            status=StageStatus.DONE,
        )
        export_dir = tmp_path / "export"
        result = run(_make_plan_dict(), state, export_dir=str(export_dir), timestamped_export=False)
        files = result.get("output_files", {})

        assert "activities_json" in files
        assert "activities_html" in files
        activities_json = Path(files["activities_json"])
        activities_html = Path(files["activities_html"])
        assert activities_json.exists()
        assert activities_html.exists()
        data = json.loads(activities_json.read_text(encoding="utf-8"))
        assert data["activities"][0]["date"] == "2026-01-17"
        assert data["activities"][0]["title"] == "Spillerutvikling JU14"
        assert "secret.example.com" not in activities_json.read_text(encoding="utf-8")
        assert "../activities.json" in activities_html.read_text(encoding="utf-8")

    def test_activity_artifacts_absent_without_activity_sheet(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_input_workbook(
            input_path,
            {
                "start_date": "2025-09-01",
                "end_date": "2025-12-01",
                "teams": [{"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"}],
                "sources": [],
            },
        )
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(
            StageName.CONFIG,
            {"input_path": str(input_path), "round_length_minutes": {}, "start_date": "2025-09-01"},
            status=StageStatus.DONE,
        )
        export_dir = tmp_path / "export"
        result = run(_make_plan_dict(), state, export_dir=str(export_dir), timestamped_export=False)
        files = result.get("output_files", {})

        assert "activities_json" not in files
        assert "activities_html" not in files
        assert not (export_dir / "activities.json").exists()
        assert not (export_dir / "activities" / "index.html").exists()

    def test_input_html_absent_and_nav_link_hidden_without_configured_input_workbook(self, tmp_path):
        """Without a Stage 1 input workbook path, input.html must not be generated and its nav link must be hidden."""
        state = PipelineState(tmp_path / "pipeline")
        export_dir = tmp_path / "export"
        result = run(_make_plan_dict(), state, export_dir=str(export_dir), timestamped_export=False)
        files = result.get("output_files", {})

        assert "input_html" not in files
        assert not (export_dir / "input.html").exists()
        html = Path(files["html"]).read_text(encoding="utf-8")
        assert "Påmeldte lag" not in html

    def test_input_html_never_exposes_internal_workbook_sheets(self, tmp_path):
        """Only the whitelisted Lag sheet may reach input.html — internal sheets must never leak."""
        input_path = tmp_path / "input.xlsx"
        _write_input_workbook(
            input_path,
            {
                "start_date": "2025-09-01",
                "end_date": "2025-12-01",
                "age_groups": ["U10"],
                "parallel_games": {"U10": 2},
                "teams": [{"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"}],
                "sources": [{"name": "Hemmelig kilde", "type": "ical", "url": "https://secret.example.com/feed.ics"}],
            },
        )
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(
            StageName.CONFIG,
            {"input_path": str(input_path), "round_length_minutes": {}},
            status=StageStatus.DONE,
        )
        export_dir = tmp_path / "export"
        result = run(_make_plan_dict(), state, export_dir=str(export_dir), timestamped_export=False)
        files = result.get("output_files", {})

        input_html = Path(files["input_html"]).read_text(encoding="utf-8")
        assert "secret.example.com" not in input_html
        assert "Hemmelig kilde" not in input_html
        assert "parallel_games" not in input_html
        assert "input.xlsx" not in input_html

    def test_conclusion_injects_blocked_count(self, tmp_path):
        """Conclusion must include blocked source count when blocked sources exist."""
        data = _make_plan_dict()
        state = PipelineState(tmp_path / "pipeline")
        # Write scraping stage with blocked sources so stage4 picks them up
        state.write_stage(
            StageName.SCRAPING,
            {"sources": [], "blocked": ["Ringerike", "Tønsberg"]},
        )
        result = run(data, state, export_dir=str(tmp_path / "export"), timestamped_export=False)
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")
        assert "2 kilde(r) blokkert." in report_html

    def test_conclusion_injects_cancelled_count(self, tmp_path):
        """Conclusion must include cancellation count when cancelled tournaments exist."""
        data = _make_plan_dict()
        # Add a cancelled tournament
        data["plan"]["tournaments"].append({
            "date": "2025-11-08",
            "arena": "Kongsberghallen",
            "age_group": "U10",
            "host_club": "Kongsberg",
            "teams": [],
            "games": [],
            "cancelled": True,
        })
        state = PipelineState(tmp_path / "pipeline")
        result = run(data, state, export_dir=str(tmp_path / "export"), timestamped_export=False)
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")
        assert "1 turnering(er) avlyst." in report_html

    def test_scrape_age_populated_from_scraping_envelope_updated_at(self, tmp_path):
        """Report HTML must show a non-empty scrape_age when the SCRAPING envelope has an updated_at field.

        This test verifies the fix for the bug where read_stage() silently dropped the
        updated_at envelope field, making scrape_age always empty in the exported report.
        The fix switches to read_envelope() so updated_at is accessible at the top level.

        Note: scrape_age renders in the metrics section of the diagnostics report page
        (html_report), not the schedule page (html), because the metrics template is only
        included when include_diagnostics=True.
        """
        state = PipelineState(tmp_path / "pipeline")
        # Write a scraping checkpoint — write_stage auto-populates updated_at in the envelope.
        state.write_stage(
            StageName.SCRAPING,
            {
                "sources": [{"name": "Kongsberg", "event_count": 5}],
                "blocked": [],
            },
            status=StageStatus.DONE,
        )
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")
        # scrape_age should be rendered in the report — any of the three time-bucket formats is valid.
        assert re.search(r"\d+[mdt] siden", report_html), (
            "Expected a scrape_age string (e.g. '5m siden', '2t siden', '1d siden') "
            "in the exported report HTML, but none was found. "
            "Check that stage4_export.py uses read_envelope (not read_stage) for the SCRAPING stage "
            "and that datetime comparison uses timezone-aware datetimes."
        )

    def test_logs_warning_when_scraping_envelope_read_fails(self, tmp_path, monkeypatch, caplog):
        state = PipelineState(tmp_path / "pipeline")
        monkeypatch.setattr(
            state,
            "read_envelope",
            lambda _stage: (_ for _ in ()).throw(RuntimeError("boom")),
        )

        with caplog.at_level(logging.WARNING, logger="tournament_scheduler.pipeline.stage4_export"):
            result = run(_make_plan_dict(), state, export_dir=str(tmp_path / "export"), timestamped_export=False)

        assert result["output_files"]["excel"]
        assert any("scraping-checkpoint" in record.message for record in caplog.records)
