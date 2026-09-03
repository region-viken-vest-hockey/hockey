from __future__ import annotations

import csv
import hashlib
import os
import subprocess
import sys
from pathlib import Path

import openpyxl

from tournament_scheduler.cli.args import build_parser
from tournament_scheduler.cli.reporting import _build_status_text
from tournament_scheduler.pipeline.state import PipelineState, StageName, StageStatus


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "rvv-miniputt"


def test_run_parser_accepts_portable_slash_command_flags() -> None:
    args = build_parser().parse_args(
        [
            "run",
            "--resume-from",
            "3",
            "--log-level",
            "verbose",
            "--force-refresh",
            "--manual-bookup-login",
            "--manual-bookup-login-timeout",
            "600",
            "--work-dir",
            ".pipeline",
        ]
    )

    assert args.command == "run"
    assert args.resume_from == "3"
    assert args.log_level == "verbose"
    assert args.force_refresh is True
    assert args.manual_bookup_login is True
    assert args.manual_bookup_login_timeout == 600


def test_run_parser_accepts_scrape_llm_flags() -> None:
    args = build_parser().parse_args(
        [
            "scrape-llm",
            "--club",
            "Holmen",
            "--work-dir",
            ".pipeline",
            "--export-dir",
            "export",
            "--endpoint",
            "http://localhost:1234",
            "--model",
            "qwen2.5-32b-instruct",
            "--max-iterations",
            "12",
            "--no-cache-results",
            "--debug-screenshots",
        ]
    )

    assert args.command == "scrape-llm"
    assert args.club == "Holmen"
    assert args.work_dir == ".pipeline"
    assert args.export_dir == "export"
    assert args.endpoint == "http://localhost:1234"
    assert args.model == "qwen2.5-32b-instruct"
    assert args.max_iterations == 12
    assert args.cache_results is False
    assert args.debug_screenshots is True


def test_repo_local_script_is_executable_and_shows_status() -> None:
    assert SCRIPT.exists()
    assert os.access(SCRIPT, os.X_OK)

    result = subprocess.run(
        [str(SCRIPT), "status"],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=True,
    )

    assert "Pipeline work-dir:" in result.stdout
    assert "Stage 1 (Config):" in result.stdout


def test_status_marks_downstream_stale_when_input_workbook_fingerprint_changes(tmp_path) -> None:
    input_file = tmp_path / "input.xlsx"
    input_file.write_bytes(b"old workbook bytes")
    old_sha = hashlib.sha256(b"old workbook bytes").hexdigest()
    work_dir = tmp_path / "pipeline"
    state = PipelineState(work_dir)
    state.write_stage(
        StageName.CONFIG,
        {
            "input_path": str(input_file),
            "input_fingerprint": {
                "algorithm": "sha256",
                "path": str(input_file),
                "sha256": old_sha,
            },
            "effective_config_fingerprint": {
                "algorithm": "sha256",
                "sha256": "old-effective",
            },
        },
        status=StageStatus.DONE,
    )
    state.write_stage(StageName.SCRAPING, {"sources": []}, status=StageStatus.DONE)
    state.write_stage(StageName.PLANNING, {"plan": {"tournaments": []}}, status=StageStatus.DONE)
    state.write_stage(StageName.EXPORT, {"output_files": {"excel": "plan.xlsx"}}, status=StageStatus.DONE)

    input_file.write_bytes(b"new workbook bytes")

    output = _build_status_text(work_dir)

    assert "Stage 2 (Scraping): failed" in output
    assert "stale from config" in output
    assert "Stage 4 (Export): failed" in output
    assert PipelineState(work_dir).is_stale(StageName.EXPORT)


def test_registrations_parser_accepts_validate_and_export_commands() -> None:
    validate_args = build_parser().parse_args(
        ["registrations", "validate", "sharepoint.csv", "--input", "input.xlsx"]
    )
    assert validate_args.command == "registrations"
    assert validate_args.registrations_command == "validate"
    assert validate_args.source == "sharepoint.csv"
    assert validate_args.input == "input.xlsx"

    export_args = build_parser().parse_args(
        [
            "registrations",
            "export",
            "sharepoint.xlsx",
            "--input",
            "input.xlsx",
            "--output",
            "input.updated.xlsx",
            "--dry-run",
        ]
    )
    assert export_args.registrations_command == "export"
    assert export_args.output == "input.updated.xlsx"
    assert export_args.dry_run is True


def test_registrations_cli_validates_and_exports_without_leaking_contact_fields(tmp_path) -> None:
    input_path = tmp_path / "input.xlsx"
    wb = openpyxl.Workbook()
    settings = wb.active
    settings.title = "Innstillinger"
    settings.append(["felt", "verdi"])
    settings.append(["start_date", "2026-10-01"])
    settings.append(["end_date", "2027-04-30"])
    age_groups = wb.create_sheet("Aldersgrupper")
    age_groups.append(["age_group", "parallel_games"])
    age_groups.append(["U10", 3])
    teams = wb.create_sheet("Lag")
    teams.append(["club", "label", "age_group"])
    teams.append(["Kongsberg", "Kongsberg 1", "U10"])
    wb.save(input_path)

    registrations = tmp_path / "sharepoint.csv"
    with open(registrations, "w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["SharePoint ID", "Klubb", "Lag", "Aldergruppe", "Status", "Kontakt", "Kommentar"],
        )
        writer.writeheader()
        writer.writerow(
            {
                "SharePoint ID": "42",
                "Klubb": "Kongsberg",
                "Lag": "Kongsberg 1",
                "Aldergruppe": "U10",
                "Status": "Godkjent",
                "Kontakt": "private@example.test",
                "Kommentar": "do not publish",
            }
        )

    validate = subprocess.run(
        [
            sys.executable,
            "-m",
            "tournament_scheduler.cli.rvv_cli",
            "registrations",
            "validate",
            str(registrations),
            "--input",
            str(input_path),
        ],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=True,
    )
    assert "Aktive lag: 1" in validate.stdout
    assert "private@example.test" not in validate.stdout
    assert "do not publish" not in validate.stdout

    output_path = tmp_path / "input.updated.xlsx"
    export = subprocess.run(
        [
            sys.executable,
            "-m",
            "tournament_scheduler.cli.rvv_cli",
            "registrations",
            "export",
            str(registrations),
            "--input",
            str(input_path),
            "--output",
            str(output_path),
        ],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=True,
    )
    assert "Skrev arbeidsbok" in export.stdout
    assert output_path.exists()
    assert output_path.with_suffix(".registrations.audit.json").exists()


def test_logs_list_subcommand_is_available_from_python_cli(tmp_path) -> None:
    log_dir = tmp_path / "export" / "2026-01-01T1200"
    log_dir.mkdir(parents=True)
    run_id = "run-20260101_120000"
    (log_dir / f"{run_id}.jsonl").write_text(
        '{"type": "run_meta", "run_id": "run-20260101_120000", '
        '"exit_status": "success", "start_time": "2026-01-01T12:00:00", '
        '"end_time": "2026-01-01T12:01:00", "duration_ms": 60000}\n',
        encoding="utf-8",
    )

    result = subprocess.run(
        [
            sys.executable, "-m", "tournament_scheduler.cli.rvv_cli",
            "logs", "list", "--count", "1", "--work-dir", str(tmp_path),
        ],
        cwd=ROOT,
        capture_output=True,
        text=True,
        check=True,
    )

    assert "Pipeline kjøringshistorie" in result.stdout
    assert "run-" in result.stdout


def test_scrape_llm_cli_points_holmen_to_deterministic_scrape() -> None:
    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "tournament_scheduler.cli.rvv_cli",
            "scrape-llm",
            "--club",
            "Holmen",
        ],
        cwd=ROOT,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 1
    assert "deterministisk skraper" in result.stdout
    assert 'rvv-miniputt scrape --club "Holmen"' in result.stdout
    assert "browser-worker" not in result.stdout
    assert "recovery-targets" in result.stdout
    assert "recovery-inject" in result.stdout


def test_scrape_llm_cli_prints_browser_tool_guidance_for_browser_only_source() -> None:
    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "tournament_scheduler.cli.rvv_cli",
            "scrape-llm",
            "--club",
            "Jutul",
        ],
        cwd=ROOT,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 1
    assert "browser-verktøy" in result.stdout
    assert "Playwright" in result.stdout
    assert "browser_worker" in result.stdout
    assert "/rvv-miniputt scrape-llm" in result.stdout
    assert "recovery-targets" in result.stdout
    assert "recovery-inject" in result.stdout
